import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from config.settings import CLASS_COLOR_MAP, LAYER_COLOR_MAP
from utils.logger import logger

class ReportGenerator:
    def __init__(self):
        self.fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    def generate_excel_report(self, df, char_df, output_path):
        """生成Excel报告"""
        try:
            wb = Workbook()
            
            # 创建明细表
            self._create_detail_sheet(wb, df)
            
            # 创建限时总览表
            self._create_summary_sheet(wb, df, char_df)
            
            # 保存文件
            wb.save(output_path)
            logger.success(f"已保存Excel报告: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"生成Excel报告失败: {e}")
            return False
    
    def _create_detail_sheet(self, wb, df):
        """创建明细表"""
        ws = wb.active
        ws.title = "明细"
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        logger.info("明细表创建完成")
    
    def _create_summary_sheet(self, wb, df, char_df):
        """创建限时总览表"""
        # 创建透视表
        pivot_df = df.pivot_table(
            index=["玩家", "角色名"],
            columns="副本",
            values="显示层数",
            aggfunc="first"
        ).fillna("-").reset_index()
        
        ws = wb.create_sheet("限时总览")
        
        # 写入数据并居中
        for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True), start=1):
            ws.append(row)
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="center", vertical="center")
        
        # 应用层数颜色标注
        self._apply_level_colors(ws, pivot_df)
        
        # 应用职业颜色标注
        self._apply_class_colors(ws, char_df)
        
        # 合并玩家列
        self._merge_player_columns(ws)
        
        logger.info("限时总览表创建完成")
    
    def _apply_level_colors(self, ws, pivot_df):
        """应用层数颜色标注"""
        for r_idx, row in enumerate(pivot_df.values, start=2):
            for c_idx, val in enumerate(row[2:], start=3):  # 跳过前两列（玩家、角色名）
                cell = ws.cell(row=r_idx, column=c_idx)
                val_str = str(cell.value)
                
                if val_str == "-":
                    cell.fill = self.fill_gray
                elif val_str.startswith("+"):
                    try:
                        level = int(re.search(r"\d+", val_str).group())
                        hex_color = LAYER_COLOR_MAP.get(level)
                        if hex_color:
                            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    except:
                        cell.fill = self.fill_gray
    
    def _apply_class_colors(self, ws, char_df):
        """应用职业颜色标注"""
        char_class_map = dict(zip(char_df["角色名"], char_df["职业"]))
        
        for row in range(2, ws.max_row + 1):
            char_name = ws.cell(row=row, column=2).value  # 第2列是角色名
            char_class = char_class_map.get(char_name)
            
            if char_class:
                hex_color = CLASS_COLOR_MAP.get(char_class)
                if hex_color:
                    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    ws.cell(row=row, column=2).fill = fill
    
    def _merge_player_columns(self, ws):
        """合并玩家列"""
        current_player = None
        start_row = 2
        
        for i in range(2, ws.max_row + 2):
            if i <= ws.max_row:
                value = ws.cell(i, 1).value  # 第1列是玩家
            else:
                value = None
            
            if value != current_player:
                if current_player is not None and i - start_row > 1:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
                    ws.cell(start_row, 1).alignment = Alignment(horizontal="center", vertical="center")
                current_player = value
                start_row = i
    
    def prepare_dataframe(self, all_records):
        """准备数据框"""
        if not all_records:
            logger.error("没有数据可生成报告")
            return None
        
        df = pd.DataFrame(all_records)
        df = df[["玩家", "角色名", "服务器", "副本", "通关时间", "限时层数", "是否限时"]]
        
        # 添加显示层数列
        df["显示层数"] = df.apply(self._format_display_level, axis=1)
        
        logger.success(f"数据框准备完成，共 {len(df)} 条记录")
        return df
    
    def _format_display_level(self, row):
        """格式化显示层数"""
        lvl = row["限时层数"]
        if pd.isna(lvl):
            return "-"
        return f"+{int(lvl)}" if row["是否限时"] == "是" else f"+{int(lvl)}*"
