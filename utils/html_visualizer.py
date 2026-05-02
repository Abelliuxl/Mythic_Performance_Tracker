import pandas as pd
import numpy as np
import json
from datetime import datetime
import traceback
from openpyxl import load_workbook
from config.settings import CLASS_COLOR_MAP, LAYER_COLOR_MAP, DUNGEON_NAME_MAP, DUNGEON_TIME_LIMIT, DUNGEON_COLOR_MAP, DUNGEON_SHORT_NAME_MAP
from utils.logger import logger

class HTMLVisualizer:
    def __init__(self):
        self.template = self._get_html_template()
    
    def generate_html_report(self, character_info_path, result_path, output_path):
        """生成HTML可视化报告"""
        try:
            # 读取数据（带兜底）
            char_df = self._safe_read_excel(character_info_path)
            result_df = self._safe_read_excel(result_path, preferred_sheet="明细")

            # 统一职业名称
            from utils.data_processor import DataProcessor
            char_df = DataProcessor.standardize_class_names(char_df)

            # 再次整体清洗，确保无数组/列表残留
            char_df = self._sanitize_dataframe(char_df)
            result_df = self._sanitize_dataframe(result_df)

            # 标准化列名：确保存在关键列
            expected_cols = {"玩家", "角色名", "服务器", "副本", "通关时间", "限时层数", "是否限时"}
            missing_cols = [c for c in ["玩家", "角色名", "副本"] if c not in result_df.columns]
            if missing_cols:
                raise ValueError(f"结果表缺少必要列: {missing_cols}")

            # 数值与标志列标准化
            result_df["限时层数"] = pd.to_numeric(result_df.get("限时层数"), errors="coerce")
            result_df["是否限时"] = result_df.get("是否限时").astype(str).str.strip()

            # 无条件重建"显示层数"列，避免原文件中携带的异常类型
            def _fmt(row):
                lvl = row.get("限时层数")
                if pd.isna(lvl):
                    return "-"
                try:
                    lvl_i = int(float(lvl))
                except Exception:
                    return "-"
                return f"+{lvl_i}" if row.get("是否限时") == "是" else f"+{lvl_i}*"
            result_df["显示层数"] = result_df.apply(_fmt, axis=1)

            # 再次清洗
            result_df = self._sanitize_dataframe(result_df)

            # 生成HTML
            html_content = self._generate_html_content(char_df, result_df)

            # 保存文件
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            logger.success(f"HTML可视化报告已生成: {output_path}")
            return True

        except Exception as e:
            logger.error(f"生成HTML报告失败: {e}\n{traceback.format_exc()}")
            return False

    def generate_html_content_only(self, character_info_path, result_path):
        """
        只生成HTML内容，不保存文件
        用于新的报告管理器架构
        """
        try:
            # 读取数据（带兜底）
            char_df = self._safe_read_excel(character_info_path)
            result_df = self._safe_read_excel(result_path, preferred_sheet="明细")

            # 统一职业名称
            from utils.data_processor import DataProcessor
            char_df = DataProcessor.standardize_class_names(char_df)

            # 再次整体清洗，确保无数组/列表残留
            char_df = self._sanitize_dataframe(char_df)
            result_df = self._sanitize_dataframe(result_df)

            # 标准化列名：确保存在关键列
            expected_cols = {"玩家", "角色名", "服务器", "副本", "通关时间", "限时层数", "是否限时"}
            missing_cols = [c for c in ["玩家", "角色名", "副本"] if c not in result_df.columns]
            if missing_cols:
                raise ValueError(f"结果表缺少必要列: {missing_cols}")

            # 数值与标志列标准化
            result_df["限时层数"] = pd.to_numeric(result_df.get("限时层数"), errors="coerce")
            result_df["是否限时"] = result_df.get("是否限时").astype(str).str.strip()

            # 无条件重建"显示层数"列，避免原文件中携带的异常类型
            def _fmt(row):
                lvl = row.get("限时层数")
                if pd.isna(lvl):
                    return "-"
                try:
                    lvl_i = int(float(lvl))
                except Exception:
                    return "-"
                return f"+{lvl_i}" if row.get("是否限时") == "是" else f"+{lvl_i}*"
            result_df["显示层数"] = result_df.apply(_fmt, axis=1)

            # 再次清洗
            result_df = self._sanitize_dataframe(result_df)

            # 生成HTML内容
            html_content = self._generate_html_content(char_df, result_df)

            logger.success("HTML内容生成成功")
            return html_content

        except Exception as e:
            logger.error(f"生成HTML内容失败: {e}\n{traceback.format_exc()}")
            return None

    def _safe_read_excel(self, file_path, preferred_sheet=None):
        """安全读取Excel：先尝试pandas，失败则退化到openpyxl逐行读取。"""
        # 优先尝试 pandas（默认参数）
        try:
            if preferred_sheet is not None:
                df = pd.read_excel(file_path, sheet_name=preferred_sheet)
            else:
                df = pd.read_excel(file_path)
            logger.info(f"读取Excel(pandas)成功: {file_path}{' / ' + preferred_sheet if preferred_sheet else ''}")
            return self._sanitize_dataframe(df)
        except Exception as e1:
            logger.warning(f"pandas读取失败，改用engine=openpyxl重试: {e1}")
            # 明确指定engine重试
            try:
                if preferred_sheet is not None:
                    df = pd.read_excel(file_path, sheet_name=preferred_sheet, engine="openpyxl")
                else:
                    df = pd.read_excel(file_path, engine="openpyxl")
                logger.info(f"读取Excel(pandas+openpyxl)成功: {file_path}{' / ' + preferred_sheet if preferred_sheet else ''}")
                return self._sanitize_dataframe(df)
            except Exception as e2:
                logger.warning(f"pandas+openpyxl仍失败，降级openpyxl逐行读取: {e2}")
                # 使用 openpyxl 读取
                wb = load_workbook(filename=file_path, data_only=True, read_only=True)
                sheet = None
                if preferred_sheet and preferred_sheet in wb.sheetnames:
                    sheet = wb[preferred_sheet]
                else:
                    sheet = wb[wb.sheetnames[0]]

                # 读取所有行并标准化
                raw_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
                if not raw_rows:
                    return pd.DataFrame()
                max_len = max(len(r) for r in raw_rows)

                def sanitize_header(val, idx):
                    if isinstance(val, (str, int, float, bool)):
                        s = str(val).strip()
                    else:
                        s = ""
                    if not s or s.lower() == 'none' or s.lower() == 'nan':
                        s = f"列{idx+1}"
                    return s

                headers = [sanitize_header(raw_rows[0][i] if i < len(raw_rows[0]) else None, i) for i in range(max_len)]
                # 去重列名
                seen = {}
                for i, name in enumerate(headers):
                    if name in seen:
                        seen[name] += 1
                        headers[i] = f"{name}_{seen[name]}"
                    else:
                        seen[name] = 1

                def coerce_scalar(v):
                    try:
                        if isinstance(v, np.ndarray):
                            return v.flatten()[0] if v.size > 0 else None
                        if isinstance(v, (list, tuple)):
                            return v[0] if len(v) > 0 else None
                        return v
                    except Exception:
                        return v

                data_dicts = []
                for r in raw_rows[1:]:
                    row_dict = {
                        headers[i]: coerce_scalar(r[i] if i < len(r) else None)
                        for i in range(max_len)
                    }
                    data_dicts.append(row_dict)

                df = pd.DataFrame(data_dicts)
                logger.info(f"读取Excel(openpyxl)成功: {file_path} / {sheet.title}")
                return self._sanitize_dataframe(df)

    def _sanitize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """将DataFrame中可能的ndarray/列表等非常规单元值压平为标量。"""
        def coerce_scalar(v):
            try:
                if isinstance(v, np.ndarray):
                    return v.flatten()[0] if v.size > 0 else None
                if isinstance(v, (list, tuple)):
                    return v[0] if len(v) > 0 else None
                return v
            except Exception:
                return v

        return df.map(coerce_scalar)
    
    def _generate_html_content(self, char_df, result_df):
        """生成HTML内容"""
        # 准备数据
        summary_data = self._prepare_summary_data(result_df, char_df)
        character_ranking_stats = self._prepare_character_ranking_stats(char_df, result_df)
        character_stats = self._prepare_character_stats(char_df, result_df)
        dungeon_stats = self._prepare_dungeon_stats(result_df)
        player_stats = self._prepare_player_stats(char_df, result_df) # 新增玩家统计数据
        character_dungeon_details = self._prepare_character_dungeon_details(result_df) # 新增角色副本详细数据
        
        # 生成图表数据
        charts_json = self._prepare_charts_data(result_df, summary_data, char_df)
        
        # 将 character_stats, CLASS_COLOR_MAP 和 player_stats 也添加到 charts_json 中，方便前端JS访问
        charts_json["character_ranking_chart_data"] = self._prepare_character_ranking_chart_data(char_df, result_df)
        charts_json["character_stats_data"] = character_stats
        charts_json["character_dungeon_details"] = character_dungeon_details # 新增
        charts_json["CLASS_COLOR_MAP"] = CLASS_COLOR_MAP
        charts_json["LAYER_COLOR_MAP"] = LAYER_COLOR_MAP # 新增层数颜色映射
        charts_json["DUNGEON_COLOR_MAP"] = DUNGEON_COLOR_MAP # 新增副本颜色映射
        charts_json["DUNGEON_FULL_NAME_MAP"] = {v: k for k, v in DUNGEON_SHORT_NAME_MAP.items()} # 新增副本全称映射，方便前端查找
        charts_json["DUNGEON_SHORT_NAME_MAP"] = DUNGEON_SHORT_NAME_MAP # 新增副本简称映射，方便前端查找
        charts_json["player_stats_data"] = player_stats # 新增玩家统计数据
        charts_json["player_character_dungeon_stats"] = self._prepare_player_character_dungeon_stats(char_df, result_df) # 新增玩家-角色-副本详细数据

        # 顶部KPI卡片
        kpi_html = self._generate_kpi_cards(result_df)
        
        # 填充模板
        html_content = self.template.replace("{{TITLE}}", "邪恶小团体大秘境统计")
        html_content = html_content.replace("{{GENERATION_TIME}}", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        html_content = html_content.replace("{{KPI_CARDS}}", kpi_html)
        html_content = html_content.replace("{{SUMMARY_TABLE}}", self._generate_summary_table(summary_data))
        html_content = html_content.replace("{{CHARACTER_RANKING}}", self._generate_character_ranking(character_ranking_stats))
        html_content = html_content.replace("{{CHARACTER_STATS}}", self._generate_character_stats(character_stats))
        html_content = html_content.replace("{{DUNGEON_STATS}}", self._generate_dungeon_stats(dungeon_stats))
        html_content = html_content.replace("{{PLAYER_STATS}}", "") # 暂时留空，后续由JS渲染

        # 将 charts_json 注入到 JavaScript 内容中
        html_content = html_content.replace("{{CHARTS_DATA}}", json.dumps(charts_json, ensure_ascii=False))

        # 注入新功能：玩家总榜 + 角色贡献环图 + 热力图增强
        player_ranking_html = self._generate_player_ranking(char_df, result_df)
        donut_html = self._generate_player_donut_charts(char_df, result_df)
        insert = f"""
    <div class="section">
        <div class="section-header">
            <h3>🏅 玩家总榜</h3>
        </div>
        {player_ranking_html}
    </div>

    <div class="section">
        <div class="section-header">
            <h3>🎯 角色贡献占比</h3>
        </div>
        <div class="donut-charts-container" id="donutChartsContainer">
            {donut_html}
        </div>
    </div>
"""
        html_content = html_content.replace("</body>", f"{insert}</body>")

        return html_content
    
    def _prepare_summary_data(self, result_df, char_df):
        """准备总览数据（避开pandas透视，提升鲁棒性）"""
        try:
            for col in ["玩家", "角色名", "副本", "显示层数"]:
                if col in result_df.columns:
                    result_df[col] = result_df[col].astype(str)

            # 创建角色到职业的映射
            char_to_class = dict(zip(char_df["角色名"], char_df["职业"]))

            matrix = {}
            for _, r in result_df.iterrows():
                player = r.get("玩家")
                char = r.get("角色名")
                dungeon = r.get("副本")
                disp = r.get("显示层数", "-")
                if pd.isna(player) or pd.isna(char) or pd.isna(dungeon):
                    continue
                key = (str(player), str(char))
                if key not in matrix:
                    matrix[key] = {}
                matrix[key].setdefault(str(dungeon), str(disp))

            summary_data = []
            for (player, char), dun_map in matrix.items():
                class_name = char_to_class.get(str(char), "未知职业")
                summary_data.append({
                    "player": player,
                    "character": char,
                    "class": class_name,
                    "dungeons": dun_map
                })
            return summary_data
        except Exception:
            logger.error("构建总览数据失败:\n" + traceback.format_exc())
            return []
    
    def _prepare_character_stats(self, char_df, result_df):
        """准备角色统计数据"""
        stats = []
        for _, char in char_df.iterrows():
            # 打印职业名称进行调试
            logger.info(f"处理角色: {char['角色名']}, 职业: '{char['职业']}'")

            char_data = result_df[result_df["角色名"] == char["角色名"]]
            # 只考虑有有效层数记录的运行
            char_data_valid_runs = char_data[pd.notna(char_data["限时层数"])]
            
            if not char_data_valid_runs.empty:
                avg_level = char_data_valid_runs["限时层数"].mean()
                timed_runs = len(char_data_valid_runs[char_data_valid_runs["是否限时"] == "是"])
                total_runs = len(char_data_valid_runs)
                
                stats.append({
                    "player": char["玩家"],
                    "character": char["角色名"],
                    "server": char["服务器"],
                    "class": char["职业"],
                    "avg_level": round(avg_level, 1),
                    "timed_runs": timed_runs,
                    "total_runs": total_runs,
                    "completion_rate": round((timed_runs / total_runs * 100), 1) if total_runs > 0 else 0, # 改为 completion_rate
                    "timed_runs_rate": round((timed_runs / total_runs * 100), 1) if total_runs > 0 else 0 # 新增限时完成率
                })
            else: # 如果该角色没有有效记录，则所有统计数据为0
                stats.append({
                    "player": char["玩家"],
                    "character": char["角色名"],
                    "server": char["服务器"],
                    "class": char["职业"],
                    "avg_level": 0,
                    "timed_runs": 0,
                    "total_runs": 0,
                    "completion_rate": 0, # 改为 completion_rate
                    "timed_runs_rate": 0 # 新增限时完成率
                })
        
        return stats

    def _prepare_character_ranking_stats(self, char_df, result_df):
        """准备角色排名数据"""
        stats = self._prepare_character_stats(char_df, result_df)
        # 按平均等级降序排序
        return sorted(stats, key=lambda x: x['avg_level'], reverse=True)

    def _prepare_character_ranking_chart_data(self, char_df, result_df):
        """为角色排名图表准备数据"""
        from config.settings import DUNGEON_COLOR_MAP, DUNGEON_TIME_LIMIT

        all_dungeon_names = list(DUNGEON_TIME_LIMIT.keys())

        # 获取所有有记录的角色
        active_characters = char_df[char_df['角色名'].isin(result_df['角色名'].unique())]

        char_dungeon_levels = {}
        for _, row in result_df.iterrows():
            char_name = row["角色名"]
            dungeon_name = row["副本"]
            level = pd.to_numeric(row["限时层数"], errors="coerce")
            if pd.notna(level):
                # 对于每个角色和副本，我们只关心最高层数
                current_max = char_dungeon_levels.get((char_name, dungeon_name), 0)
                char_dungeon_levels[(char_name, dungeon_name)] = max(current_max, level)

        char_scores = {}
        for char_name in active_characters['角色名']:
            score = sum(char_dungeon_levels.get((char_name, dn), 0) for dn in all_dungeon_names)
            char_scores[char_name] = score

        # 按分数排序
        sorted_chars = sorted(char_scores.keys(), key=lambda c: char_scores[c], reverse=True)

        datasets = []
        for dungeon_name in all_dungeon_names:
            dataset_data = []
            for char_name in sorted_chars:
                level = char_dungeon_levels.get((char_name, dungeon_name), 0)
                dataset_data.append(level)
            
            datasets.append({
                "label": dungeon_name,
                "backgroundColor": DUNGEON_COLOR_MAP.get(dungeon_name, "rgba(120, 120, 120, 0.8)"),
                "borderColor": DUNGEON_COLOR_MAP.get(dungeon_name, "rgba(120, 120, 120, 1)").replace("0.8)", "1)"),
                "borderWidth": 1,
                "data": dataset_data,
            })

        # 获取角色到职业的映射
        char_to_class = dict(zip(char_df["角色名"], char_df["职业"]))
        char_classes = [char_to_class.get(char, "未知职业") for char in sorted_chars]

        return {
            "labels": sorted_chars,
            "datasets": datasets,
            "classes": char_classes
        }
    
    def _prepare_player_stats(self, char_df, result_df):
        """准备玩家统计数据，用于堆叠柱状图"""
        from config.settings import DUNGEON_COLOR_MAP, DUNGEON_TIME_LIMIT

        all_dungeon_names = list(DUNGEON_TIME_LIMIT.keys()) # All 8 dungeons

        # Step 1: Prepare character-dungeon level mapping, filling missing dungeons with 0
        # { (character_name, dungeon_name): level }
        char_dungeon_levels = {}
        for _, row in result_df.iterrows():
            char_name = row["角色名"]
            dungeon_name = row["副本"]
            level = pd.to_numeric(row["限时层数"], errors="coerce")
            if pd.isna(level):
                level = 0 # Treat NaN levels as 0 for calculation purposes
            char_dungeon_levels[(char_name, dungeon_name)] = level

        # Identify characters that have at least one record with a non-zero level in result_df
        active_characters = set(
            result_df[pd.to_numeric(result_df["限时层数"], errors='coerce').fillna(0) > 0]["角色名"].unique()
        )
        
        # Filter char_df to only include active characters
        active_char_df = char_df[char_df["角色名"].isin(active_characters)]
        
        # Group active characters by player
        player_char_map = active_char_df.groupby("玩家")["角色名"].apply(list).to_dict()

        # Step 2: Determine player_dungeon_max_levels
        # {player_name: {dungeon_name: max_level_for_this_dungeon_across_all_player_chars}}
        player_dungeon_max_levels = {}
        player_dungeon_actual_runs = {} # To store actual runs for tooltip

        for player_name, characters in player_char_map.items():
            player_dungeon_max_levels[player_name] = {}
            player_dungeon_actual_runs[player_name] = {}

            for dungeon_name in all_dungeon_names:
                max_level_for_dungeon = 0
                total_runs_for_dungeon = 0 # Actual runs for this dungeon across all player's chars

                for char_name in characters:
                    level = char_dungeon_levels.get((char_name, dungeon_name), 0)
                    max_level_for_dungeon = max(max_level_for_dungeon, level)
                    
                    # Count actual runs for tooltip
                    if (char_name, dungeon_name) in char_dungeon_levels and char_dungeon_levels[(char_name, dungeon_name)] > 0:
                        total_runs_for_dungeon += 1
                
                player_dungeon_max_levels[player_name][dungeon_name] = max_level_for_dungeon
                player_dungeon_actual_runs[player_name][dungeon_name] = total_runs_for_dungeon

        # Step 3: Calculate player_overall_stats based on player_dungeon_max_levels
        player_overall_stats = {}
        for player_name, dungeon_levels in player_dungeon_max_levels.items():
            total_level_sum_for_player = sum(dungeon_levels.values()) # Sum of 8 highest levels
            player_overall_stats[player_name] = {
                "total_level_sum": total_level_sum_for_player,
                "total_potential_runs": len(all_dungeon_names) # Always 8 dungeons for the average
            }
        
        # Calculate average level and sort players
        sorted_players = sorted(
            player_overall_stats.keys(),
            key=lambda p: player_overall_stats[p]["total_level_sum"] / player_overall_stats[p]["total_potential_runs"],
            reverse=True
        )
        
        # Step 4: Build Chart.js datasets
        datasets = []
        for dungeon_name in all_dungeon_names:
            dataset_data = []
            dungeon_meta_avg_levels = [] # For tooltip (highest level for this dungeon)
            dungeon_meta_runs = [] # For tooltip (actual runs for this dungeon)

            for player_name in sorted_players:
                max_level_for_dungeon = player_dungeon_max_levels[player_name][dungeon_name]
                actual_runs_for_dungeon = player_dungeon_actual_runs[player_name][dungeon_name]
                
                # Contribution to the player's overall average (max_level / 8)
                contribution = round(max_level_for_dungeon / len(all_dungeon_names), 2)
                
                dataset_data.append(contribution)
                dungeon_meta_avg_levels.append(max_level_for_dungeon)
                dungeon_meta_runs.append(actual_runs_for_dungeon)
            
            datasets.append({
                "label": dungeon_name,
                "backgroundColor": DUNGEON_COLOR_MAP.get(dungeon_name, "rgba(120, 120, 120, 0.8)"),
                "borderColor": DUNGEON_COLOR_MAP.get(dungeon_name, "rgba(120, 120, 120, 1)").replace("0.8)", "1)"),
                "borderWidth": 1,
                "data": dataset_data,
                "meta": {
                    "avg_levels": dungeon_meta_avg_levels, # Now stores max level for this dungeon
                    "runs": dungeon_meta_runs # Now stores actual runs for this dungeon
                }
            })
        
        return {
            "player_labels": sorted_players,
            "datasets": datasets
        }

    def _prepare_dungeon_stats(self, result_df):
        """准备副本统计数据"""
        stats = []
        for dungeon_full_name in result_df["副本"].unique():
            dungeon_short_name = DUNGEON_SHORT_NAME_MAP.get(dungeon_full_name, dungeon_full_name) # 获取简称
            dungeon_data = result_df[result_df["副本"] == dungeon_full_name]
            # 只考虑有有效层数记录的运行
            dungeon_data_valid_runs = dungeon_data[pd.notna(dungeon_data["限时层数"])]

            avg_time = self._time_to_seconds(dungeon_data_valid_runs["通关时间"]).mean() if not dungeon_data_valid_runs.empty else 0
            avg_level = dungeon_data_valid_runs["限时层数"].mean() if not dungeon_data_valid_runs.empty else 0
            timed_runs = len(dungeon_data_valid_runs[dungeon_data_valid_runs["是否限时"] == "是"])
            total_runs = len(dungeon_data_valid_runs)
            
            stats.append({
                "dungeon_full_name": dungeon_full_name, # 存储全称
                "dungeon": dungeon_full_name, # 副本统计板块使用全称
                "avg_time": self._seconds_to_time_format(avg_time),
                "avg_level": round(avg_level, 1),
                "timed_runs": timed_runs,
                "total_runs": total_runs,
                "completion_rate": round((timed_runs / total_runs * 100), 1) if total_runs > 0 else 0 # 改为 completion_rate
            })
        
        return stats
    
    def _generate_dungeon_stats(self, dungeon_stats):
        """生成副本统计HTML"""
        # from config.settings import DUNGEON_COLOR_MAP # 导入副本颜色映射 - 已经全局导入了

        html = """
        <div class="stats-grid">
            <h3>🗺️ 副本统计</h3>
            <div class="stats-cards">
        """
        
        for stat in dungeon_stats:
            dungeon_color = DUNGEON_COLOR_MAP.get(stat["dungeon_full_name"], "rgba(120, 120, 120, 0.8)") # 获取副本颜色
            html += f"""
                <div class="stat-card">
                    <div class="card-header" style="background-color: {dungeon_color};">
                        <div class="dungeon-name">{stat["dungeon_full_name"]}</div>
                    </div>
                    <div class="card-content">
                        <div class="stat-item">
                            <span class="stat-label">平均等级</span>
                            <span class="stat-value">{stat["avg_level"]}</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-label">平均时间</span>
                            <span class="stat-value">{stat["avg_time"]}</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-label">通关率</span>
                            <span class="stat-value">{stat["completion_rate"]}%</span>
                        </div>
                    </div>
                </div>
            """
        
        html += """
            </div>
        </div>
        """
        
        return html
    
    def _get_level_class(self, level):
        """获取等级样式类"""
        if level == "-":
            return "level-empty"
        try:
            level_num = int(level.replace("+", "").replace("*", ""))
            return f"level-{level_num}"
        except:
            return "level-empty"
    
    def _time_to_seconds(self, time_series):
        """将时间字符串转换为秒数"""
        seconds = []
        for time_str in time_series:
            try:
                parts = time_str.split(":")
                seconds.append(int(parts[0]) * 60 + int(parts[1]))
            except:
                seconds.append(0)
        return pd.Series(seconds)
    
    def _seconds_to_time_format(self, seconds):
        """将秒数转换为时间格式"""
        minutes = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{minutes:02d}:{secs:02d}"

    def _prepare_player_character_dungeon_stats(self, char_df, result_df):
        """
        准备玩家-角色-副本详细数据，用于前端弹窗显示。
        结构: {player_name: [{character_info, dungeon_stats: {dungeon_name: {avg_level, timed_runs, total_runs}}}]}
        """
        player_char_dungeon_stats = {}
        
        # 构建角色到玩家、职业、服务器的映射
        char_to_info = {}
        for _, row in char_df.iterrows():
            char_to_info[row["角色名"]] = {
                "player": row["玩家"],
                "class": row["职业"],
                "server": row["服务器"]
            }

        # 遍历结果数据，聚合每个角色在每个副本的统计
        # 使用字典来存储中间聚合结果，键为 (角色名, 副本名)
        char_dungeon_agg = {} # {(char, dungeon): {"level_sum": X, "timed_runs": Y, "total_runs": Z}}

        for _, row in result_df.iterrows():
            char_name = row["角色名"]
            dungeon_name = row["副本"]
            level = pd.to_numeric(row["限时层数"], errors="coerce")
            is_timed = (row["是否限时"] == "是")

            if pd.isna(level) or char_name not in char_to_info:
                continue # 跳过无效记录或未在角色信息中找到的角色

            key = (char_name, dungeon_name)
            if key not in char_dungeon_agg:
                char_dungeon_agg[key] = {"level_sum": 0, "timed_runs": 0, "total_runs": 0}
            
            char_dungeon_agg[key]["level_sum"] += level
            char_dungeon_agg[key]["total_runs"] += 1
            if is_timed:
                char_dungeon_agg[key]["timed_runs"] += 1
        
        # 将聚合结果组织成最终需要的结构
        for (char_name, dungeon_name), stats in char_dungeon_agg.items():
            char_info = char_to_info.get(char_name)
            if not char_info:
                continue # 再次检查，确保角色信息存在

            player_name = char_info["player"]
            
            if player_name not in player_char_dungeon_stats:
                player_char_dungeon_stats[player_name] = []

            # 查找该角色是否已存在于玩家列表中
            found_char = None
            for existing_char_entry in player_char_dungeon_stats[player_name]:
                if existing_char_entry["character"] == char_name:
                    found_char = existing_char_entry
                    break
            
            if not found_char:
                found_char = {
                    "character": char_name,
                    "class": char_info["class"],
                    "server": char_info["server"],
                    "dungeon_stats": {}
                }
                player_char_dungeon_stats[player_name].append(found_char)
            
            # 添加副本统计数据
            found_char["dungeon_stats"][dungeon_name] = {
                "avg_level": round(stats["level_sum"] / stats["total_runs"], 1) if stats["total_runs"] > 0 else 0,
                "timed_runs": stats["timed_runs"],
                "total_runs": stats["total_runs"]
            }
        
        return player_char_dungeon_stats

    def _prepare_character_dungeon_details(self, result_df):
        """
        准备每个角色在各个副本的详细数据，用于弹窗。
        返回: { "character_key": { "dungeon_name": { "timed_runs": X, "total_runs": Y, "avg_level": Z, "completion_rate": P } } }
        """
        char_dungeon_details = {}
        
        # 确保 '限时层数' 是数字类型
        result_df['限时层数'] = pd.to_numeric(result_df['限时层数'], errors='coerce')

        # 按角色和副本分组
        # 使用 '角色名' 和 '服务器' 来创建唯一键
        grouped = result_df.groupby(['角色名', '服务器', '副本'])

        for (char_name, server, dungeon_name), group in grouped:
            key = f"{char_name}-{server}"
            if key not in char_dungeon_details:
                char_dungeon_details[key] = {}

            valid_runs = group.dropna(subset=['限时层数'])
            if valid_runs.empty:
                continue

            timed_runs = valid_runs[valid_runs['是否限时'] == '是'].shape[0]
            total_runs = valid_runs.shape[0]
            avg_level = valid_runs['限时层数'].mean()
            completion_rate = round((timed_runs / total_runs * 100), 1) if total_runs > 0 else 0

            char_dungeon_details[key][dungeon_name] = {
                "timed_runs": timed_runs,
                "total_runs": total_runs,
                "avg_level": round(avg_level, 2),
                "completion_rate": completion_rate
            }
            
        return char_dungeon_details

    def _generate_player_ranking(self, char_df, result_df):
        from config.settings import DUNGEON_TIME_LIMIT

        all_dungeons = list(DUNGEON_TIME_LIMIT.keys())
        player_chars = {}
        for _, row in char_df.iterrows():
            p = row["玩家"]
            if p not in player_chars:
                player_chars[p] = []
            player_chars[p].append(row["角色名"])

        labels, data, char_counts, top_scores = [], [], [], []
        for player, chars in player_chars.items():
            total = 0
            char_scores = []
            for cname in chars:
                cdf = result_df[result_df["角色名"] == cname]
                score = 0
                for dn in all_dungeons:
                    match = cdf[cdf["副本"] == dn]
                    if not match.empty:
                        lvl = pd.to_numeric(match["限时层数"].iloc[0], errors="coerce")
                        if pd.notna(lvl):
                            score += int(lvl)
                char_scores.append(score)
                total += score
            top = max(char_scores) if char_scores else 0
            labels.append(player)
            data.append(total)
            char_counts.append(len(chars))
            top_scores.append(top)

        import json
        return f"""
<div class="chart-card">
    <div class="chart-container">
        <canvas id="playerRankingChart"></canvas>
    </div>
</div>
<script>
new Chart(document.getElementById('playerRankingChart'), {{
    type: 'bar',
    data: {{
        labels: {json.dumps(labels)},
        datasets: [{{
            label: '总分',
            data: {json.dumps(data)},
            backgroundColor: {json.dumps(["#C41F3B" if c > 1 else "#69CCF0" for c in char_counts])},
            borderColor: '#000000',
            borderWidth: 1,
        }}]
    }},
    options: {{
        responsive: true,
        maintainAspectRatio: false,
        indexAxis: 'y',
        plugins: {{
            legend: {{ display: false }},
            tooltip: {{
                callbacks: {{
                    afterLabel: function(ctx) {{
                        var i = ctx.dataIndex;
                        return '角色数: ' + {json.dumps(char_counts)}[i] + ' | 最高角色: ' + {json.dumps(top_scores)}[i] + '分';
                    }}
                }}
            }}
        }},
        scales: {{
            x: {{ beginAtZero: true, title: {{ display: true, text: '总分' }} }},
            y: {{ title: {{ display: true, text: '' }} }}
        }}
    }}
}});
</script>"""

    def _generate_player_donut_charts(self, char_df, result_df):
        from config.settings import CLASS_COLOR_MAP, DUNGEON_TIME_LIMIT

        all_dungeons = list(DUNGEON_TIME_LIMIT.keys())
        player_chars = {}
        for _, row in char_df.iterrows():
            p = row["玩家"]
            if p not in player_chars:
                player_chars[p] = []
            player_chars[p].append((row["角色名"], row["职业"]))

        idx = 0
        html = ""
        for player, chars in player_chars.items():
            if len(chars) < 2:
                continue
            labels, data, colors = [], [], []
            total_score = 0
            for cname, cclass in chars:
                cdf = result_df[result_df["角色名"] == cname]
                score = 0
                for dn in all_dungeons:
                    match = cdf[cdf["副本"] == dn]
                    if not match.empty:
                        lvl = pd.to_numeric(match["限时层数"].iloc[0], errors="coerce")
                        if pd.notna(lvl):
                            score += int(lvl)
                labels.append(cname)
                data.append(score)
                hc = CLASS_COLOR_MAP.get(cclass, "888888")
                colors.append(f"#{hc}")
                total_score += score

            if total_score == 0:
                continue

            canvas_id = f"donutChart{idx}"
            html += f"""
        <div class="chart-card donut-chart-card">
            <h4>{player}</h4>
            <div class="chart-container-donut">
                <canvas id="{canvas_id}"></canvas>
            </div>
            <script>
                new Chart(document.getElementById('{canvas_id}'), {{
                    type: 'doughnut',
                    data: {{
                        labels: {json.dumps(labels)},
                        datasets: [{{
                            data: {json.dumps(data)},
                            backgroundColor: {json.dumps(colors)},
                            borderColor: '#000000',
                            borderWidth: 1,
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            legend: {{ position: 'bottom', labels: {{ font: {{ size: 12 }}, boxWidth: 12 }} }},
                            tooltip: {{
                                callbacks: {{
                                    label: function(ctx) {{
                                        let pct = ((ctx.raw / {total_score}) * 100).toFixed(1);
                                        return ctx.label + ': ' + ctx.raw + '分 (' + pct + '%)';
                                    }}
                                }}
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
"""
            idx += 1

        if not html:
            return "<p style='padding:20px;color:#888;'>没有多角色玩家，无数据显示</p>"
        return f'<div class="charts-container donut-grid">{html}</div>'

    def _get_html_template(self):
        """获取HTML模板"""
        template_path = "utils/templates/report_template.html"
        css_path = "utils/static/css/report_style.css"
        js_path = "utils/static/js/report_script.js"

        with open(template_path, 'r', encoding='utf-8') as f:
            template_content = f.read()
        
        with open(css_path, 'r', encoding='utf-8') as f:
            css_content = f.read()
            # 动态添加LAYER_COLOR_MAP的样式
            layer_color_styles = "\n".join([f"""
        .level-{level} {{
            background-color: #{color} !important;
            color: #1f2937;
            font-weight: bold;
            border: 1px solid #000000;
        }}
        """ for level, color in LAYER_COLOR_MAP.items()])
            css_content = css_content + layer_color_styles

        with open(js_path, 'r', encoding='utf-8') as f:
            js_content = f.read()
        
        # 将CSS和JS内容注入到模板中
        template_content = template_content.replace("/* CSS will be injected here */", css_content)
        template_content = template_content.replace("/* JavaScript will be injected here */", js_content)
        
        return template_content
    
    def _prepare_charts_data(self, result_df, summary_data, char_df):
        """准备图表数据"""
        charts = {
            "level_distribution": {},
            "class_performance": {},
            "dungeon_performance": {},
            "player_performance": {} # 新增玩家性能数据
        }
        
        # 等级分布（只统计有效数值）
        levels_series = pd.to_numeric(result_df["限时层数"], errors="coerce").dropna()
        level_counts = levels_series.value_counts().sort_index()
        charts["level_distribution"] = {
            "labels": [f"+{int(l)}" for l in level_counts.index.tolist()],
            "data": [int(x) for x in level_counts.values.tolist()]
        }
        
        # 职业表现（按职业聚合）
        char_to_class = dict(zip(char_df["角色名"], char_df["职业"]))
        class_levels = {}
        for _, row in result_df.iterrows():
            cls = char_to_class.get(row["角色名"], "未知职业")
            lvl = row["限时层数"]
            if pd.notna(lvl):
                class_levels.setdefault(cls, []).append(float(lvl))

        for cls, levels in class_levels.items():
            charts["class_performance"][cls] = {
                "avg_level": round(sum(levels) / len(levels), 1),
                "count": len(levels),
                "color": f"#{CLASS_COLOR_MAP.get(cls, '888888')}"
            }
        
        # 副本表现
        dungeon_labels = [] # 简称
        dungeon_full_names = [] # 全称
        dungeon_avg_levels = []
        dungeon_timed_rates = []
        
        for dungeon_full_name in result_df["副本"].dropna().unique():
            dungeon_short_name = DUNGEON_SHORT_NAME_MAP.get(dungeon_full_name, dungeon_full_name) # 获取简称
            dungeon_data = result_df[result_df["副本"] == dungeon_full_name]
            avg_lvl = pd.to_numeric(dungeon_data["限时层数"], errors="coerce").mean()
            timed_rate = 0
            if len(dungeon_data) > 0:
                timed_rate = round((dungeon_data["是否限时"].astype(str).str.strip() == "是").sum() / len(dungeon_data) * 100, 1)
            
            dungeon_labels.append(str(dungeon_short_name))
            dungeon_full_names.append(str(dungeon_full_name)) # 保存全称
            dungeon_avg_levels.append(round(avg_lvl, 1) if pd.notna(avg_lvl) else 0)
            dungeon_timed_rates.append(timed_rate)
            
        charts["dungeon_performance"] = {
            "labels": dungeon_labels,
            "full_names": dungeon_full_names, # 新增全称
            "avg_levels": dungeon_avg_levels,
            "timed_rates": dungeon_timed_rates
        }
        
        return charts

    def _generate_kpi_cards(self, result_df):
        """生成顶部关键指标KPI卡片"""
        # 只考虑有有效层数记录的运行
        result_df_valid_runs = result_df[pd.notna(result_df["限时层数"])]

        total_runs = len(result_df_valid_runs)
        total_chars = result_df_valid_runs["角色名"].nunique() if not result_df_valid_runs.empty else 0
        timed_runs = (result_df_valid_runs["是否限时"] == "是").sum()
        completion_rate = round((timed_runs / total_runs * 100), 1) if total_runs > 0 else 0 # 改为 completion_rate
        avg_level = round(pd.to_numeric(result_df_valid_runs["限时层数"], errors='coerce').mean(), 1) if not result_df_valid_runs.empty else 0

        return f"""
        <div class=\"kpi-grid\">
            <div class=\"kpi-card\">
                <div class=\"kpi-label\">📊 总运行数</div>
                <div class=\"kpi-value\">{total_runs}</div>
            </div>
            <div class=\"kpi-card\">
                <div class=\"kpi-label\">👥 角色数量</div>
                <div class=\"kpi-value\">{total_chars}</div>
            </div>
            <div class=\"kpi-card\">
                <div class=\"kpi-label\">⏱️ 8本通关率</div>
                <div class=\"kpi-value\">{completion_rate}%</div>
            </div>
            <div class=\"kpi-card\">
                <div class=\"kpi-label\">📈 平均层数</div>
                <div class=\"kpi-value\">{avg_level}</div>
            </div>
        </div>
        """
    
    def _hex_to_rgba(self, hex_color, alpha=0.1):
        """将十六进制颜色转换为RGBA字符串"""
        hex_color = hex_color.lstrip('#')
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return f"rgba({r}, {g}, {b}, {alpha})"

    def _generate_summary_table(self, summary_data):
        """生成总览表格HTML（搜索 + 冻结前两列 + 可排序）"""
        html = """
        <div class=\"table-container\">
            <div class=\"table-toolbar\">
                <div class=\"toolbar-actions\">
                    <input id=\"summarySearch\" class=\"search-input\" type=\"search\" placeholder=\"搜索玩家/角色/副本...\" />
                    <button id=\"clearSearch\" class=\"btn\">清除</button>
                </div>
            </div>
            <div class=\"table-wrapper\">
                <table id=\"summaryTable\" class=\"summary-table\">
                    <thead>
                        <tr>
                            <th class=\"sticky-col sticky-col-1 sortable\" data-type=\"text\">👤 玩家</th>
                            <th class=\"sticky-col sticky-col-2 sortable\" data-type=\"text\">🎮 角色名</th>
        """

        # 添加副本列头（先按配置顺序，再追加未配置副本；使用全量union）
        dungeons = []
        if summary_data:
            seen = set()
            union = []
            for item in summary_data:
                for d in item.get("dungeons", {}).keys():
                    if d not in seen:
                        seen.add(d)
                        union.append(d)
            preferred = list(DUNGEON_TIME_LIMIT.keys())
            dungeons = [d for d in preferred if d in seen] + [d for d in union if d not in preferred]
            for dungeon_full_name in dungeons:
                dungeon_short_name = DUNGEON_SHORT_NAME_MAP.get(dungeon_full_name, dungeon_full_name)
                html += f'<th class="sortable" data-type="level" title="{dungeon_full_name}">{dungeon_short_name}</th>'

        html += """
                        </tr>
                    </thead>
                    <tbody>
        """

        # 添加数据行
        for player_data in summary_data:
            class_color = CLASS_COLOR_MAP.get(player_data["class"], "FFFFFF")
            rgba_color = self._hex_to_rgba(class_color, 0.1)
            html += f"""
                        <tr>
                            <td class="sticky-col sticky-col-1">{player_data["player"]}</td>
                            <td class="sticky-col sticky-col-2" style="background-color: {rgba_color}; border: 1px solid #000000;">{player_data["character"]}</td>
            """
            for dungeon_full_name in dungeons:
                level = player_data["dungeons"].get(dungeon_full_name, "-")
                level_class = self._get_level_class(level)
                html += f'<td class="{level_class}" title="{dungeon_full_name}">{level}</td>'
            html += "</tr>"

        html += """
                    </tbody>
                </table>
            </div>
        </div>
        """

        return html
    
    def _generate_character_stats(self, character_stats):
        """生成角色统计HTML"""
        html = """
        <div class="stats-grid">
            <h3>🧙 角色统计</h3>
            <div class="stats-cards">
        """
        
        for stat in character_stats:
            class_color = CLASS_COLOR_MAP.get(stat["class"], "FFFFFF")
            character_key = f"{stat['character']}-{stat['server']}"
            html += f"""
                <div class="stat-card character-stat-card" data-character-key="{character_key}">
                    <div class="card-header" style="border-left-color: #{class_color};">
                        <div class="character-info">
                            <div class="character-name">{stat["character"]}</div>
                            <div class="character-server">{stat["server"]}</div>
                        </div>
                        <div class="character-class" style="color: #{class_color};">{stat["class"]}</div>
                    </div>
                    <div class="card-content">
                        <div class="stat-item">
                            <span class="stat-label">平均等级</span>
                            <span class="stat-value">{stat["avg_level"]}</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-label">限时完成</span>
                            <span class="stat-value">{stat["timed_runs"]}/{stat["total_runs"]}</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-label">通关率</span>
                            <span class="stat-value">{stat["completion_rate"]}%</span>
                        </div>
                    </div>
                </div>
            """
        
        html += """
            </div>
        </div>
        """
        
        return html

    def _generate_character_ranking(self, character_ranking_stats):
        """生成角色排名HTML"""
        return """
        <div class="section-header">
            <h3>🏆 角色排名</h3>
        </div>
        <div class="charts-container">
            <div class="chart-card">
                <h4>👑 角色表现</h4>
                <div class="chart-container">
                    <canvas id="characterRankingChart"></canvas>
                </div>
            </div>
        </div>
        """
    
    def _generate_dungeon_stats(self, dungeon_stats):
        """生成副本统计HTML"""
        # from config.settings import DUNGEON_COLOR_MAP # 导入副本颜色映射 - 已经全局导入了

        html = """
        <div class="stats-grid">
            <h3>🗺️ 副本统计</h3>
            <div class="stats-cards">
        """
        
        for stat in dungeon_stats:
            dungeon_color = DUNGEON_COLOR_MAP.get(stat["dungeon_full_name"], "rgba(120, 120, 120, 0.8)") # 获取副本颜色
            html += f"""
                <div class="stat-card">
                    <div class="card-header" style="background-color: {dungeon_color};">
                        <div class="dungeon-name">{stat["dungeon_full_name"]}</div>
                    </div>
                    <div class="card-content">
                        <div class="stat-item">
                            <span class="stat-label">平均等级</span>
                            <span class="stat-value">{stat["avg_level"]}</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-label">平均时间</span>
                            <span class="stat-value">{stat["avg_time"]}</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-label">通关率</span>
                            <span class="stat-value">{stat["completion_rate"]}%</span>
                        </div>
                    </div>
                </div>
            """
        
        html += """
            </div>
        </div>
        """
        
        return html
    
    def _get_level_class(self, level):
        """获取等级样式类"""
        if level == "-":
            return "level-empty"
        try:
            level_num = int(level.replace("+", "").replace("*", ""))
            return f"level-{level_num}"
        except:
            return "level-empty"
    
    def _time_to_seconds(self, time_series):
        """将时间字符串转换为秒数"""
        seconds = []
        for time_str in time_series:
            try:
                parts = time_str.split(":")
                seconds.append(int(parts[0]) * 60 + int(parts[1]))
            except:
                seconds.append(0)
        return pd.Series(seconds)
    
    def _seconds_to_time_format(self, seconds):
        """将秒数转换为时间格式"""
        minutes = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{minutes:02d}:{secs:02d}"
    
    def _get_html_template(self):
        """获取HTML模板"""
        template_path = "utils/templates/report_template.html"
        css_path = "utils/static/css/report_style.css"
        js_path = "utils/static/js/report_script.js"

        with open(template_path, 'r', encoding='utf-8') as f:
            template_content = f.read()
        
        with open(css_path, 'r', encoding='utf-8') as f:
            css_content = f.read()
            # 动态添加LAYER_COLOR_MAP的样式
            layer_color_styles = "\n".join([f"""
        .level-{level} {{
            background-color: #{color} !important;
            color: #1f2937;
            font-weight: bold;
            border: 1px solid #000000;
        }}
        """ for level, color in LAYER_COLOR_MAP.items()])
            css_content = css_content + layer_color_styles

        with open(js_path, 'r', encoding='utf-8') as f:
            js_content = f.read()
        
        # 将CSS和JS内容注入到模板中
        template_content = template_content.replace("/* CSS will be injected here */", css_content)
        template_content = template_content.replace("/* JavaScript will be injected here */", js_content)
        
        return template_content
