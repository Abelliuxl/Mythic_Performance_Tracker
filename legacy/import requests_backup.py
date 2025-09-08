import pandas as pd
import re
import time
from urllib.parse import quote
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ===== 副本映射与限时阈值（秒） =====
dungeon_name_map = {
    "Cinderbrew Meadery": "燧酿酒庄",
    "Operation: Mechagon - Workshop": "麦卡贡行动 - 车间",
    "Operation: Floodgate": "水闸行动",
    "Darkflame Cleft": "暗焰裂口",
    "The Rookery": "驭雷栖巢",
    "The MOTHERLODE!!": "暴富矿区！！",
    "Theater of Pain": "伤逝剧场",
    "Priory of the Sacred Flame": "圣焰隐修院",
}

dungeon_time_limit = {
    "燧酿酒庄": 33 * 60,
    "伤逝剧场": 34 * 60,
    "圣焰隐修院": 32 * 60 + 30,
    "暗焰裂口": 31 * 60,
    "暴富矿区！！": 33 * 60,
    "水闸行动": 33 * 60,
    "驭雷栖巢": 29 * 60,
    "麦卡贡行动 - 车间": 32 * 60,
}

# ===== 日志系统 =====
log_entries = []
def log(msg):
    print(msg)
    log_entries.append(msg)

# ===== 浏览器配置 =====
def create_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument("--proxy-server=http://127.0.0.1:7890")
    service = Service(executable_path="D:/ducument/chromedriver-win64/chromedriver.exe")
    return webdriver.Chrome(service=service, options=options)

# ===== 构建 URL =====
def build_url(server, character_name):
    return f"https://www.warcraftlogs.com/character/cn/{quote(server.strip(), safe='')}/{quote(character_name.strip(), safe='')}?zone=43"

# ===== 单角色爬取 =====
def scrape_character(driver, server, character_name):
    url = build_url(server, character_name)
    driver.get(url)

    try:
        # ✅ 最多等 12 秒，直到副本表格加载出来
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "tr[role='row']"))
        )
    except:
        log(f"[超时] 页面未能加载出副本表格：{url}")
        return []

    soup = BeautifulSoup(driver.page_source, 'html.parser')
    rows = soup.find_all("tr", {"role": "row"})

    if not rows:
        log(f"[警告] 无法找到副本表格，页面可能未加载或角色无记录：{url}")
        return []

    records = []
    for row in rows:
        try:
            a_tag = row.find("a", class_="Boss zone-boss-cell")
            if not a_tag:
                continue
            raw_dungeon = a_tag.text.strip()
            dungeon = dungeon_name_map.get(raw_dungeon, raw_dungeon)

            time_cells = row.find_all("td", class_="verbose main-table-number kills-cell")
            if len(time_cells) < 2:
                continue

            time_text = time_cells[1].get_text(strip=True)
            time_match = re.search(r"\d{1,2}:\d{2}", time_text)
            time_str = time_match.group(0) if time_match else "未知"
            level_match = re.search(r"\+(\d+)", time_text)
            plus_level = int(level_match.group(1)) if level_match else None

            t_split = time_str.split(":")
            run_seconds = int(t_split[0]) * 60 + int(t_split[1]) if len(t_split) == 2 else 9999
            limit = dungeon_time_limit.get(dungeon)
            on_time = (limit is not None and run_seconds <= limit)
            result = "是" if on_time else "否"

            records.append({
                "副本": dungeon,
                "限时层数": plus_level,
                "通关时间": time_str,
                "是否限时": result
            })
        except Exception as e:
            log(f"[错误] 解析副本行失败：{e}")
            continue

    return records

# ===== 主程序入口 =====
try:
    char_df = pd.read_excel("character_info.xlsx")
    log("✅ 成功读取 character_info.xlsx")
except Exception as e:
    log(f"❌ 无法读取角色文件：{e}")
    exit()

all_records = []

for _, row in char_df.iterrows():
    player = str(row["玩家"]).strip()
    name = str(row["角色名"]).strip()
    server = str(row["服务器"]).strip()

    log(f"\n—— 开始抓取：{player} / {name}（{server}）")
    driver = create_driver()
    try:
        char_data = scrape_character(driver, server, name)
        log(f"→ 获取成功，共 {len(char_data)} 条记录")
        for entry in char_data:
            entry.update({"玩家": player, "角色名": name, "服务器": server})
            all_records.append(entry)
    except Exception as e:
        log(f"⚠ 抓取失败：{name}：{e}")
    finally:
        driver.quit()

if not all_records:
    log("❌ 没有任何副本数据被抓取，终止写入。")
else:
    df = pd.DataFrame(all_records)
    df = df[["玩家", "角色名", "服务器", "副本", "通关时间", "限时层数", "是否限时"]]

    # 添加格式化列
    def format_level(row):
        lvl = row["限时层数"]
        if pd.isna(lvl):
            return "-"
        return f"+{int(lvl)}" if row["是否限时"] == "是" else f"+{int(lvl)}*"
    df["显示层数"] = df.apply(format_level, axis=1)

    pivot_df = df.pivot_table(
        index=["玩家", "角色名"],
        columns="副本",
        values="显示层数",
        aggfunc="first"
    ).fillna("-")

    # Excel 输出并着色
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "明细"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("限时总览")

    fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill_red  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=True, header=True)):
        ws2.append(row)
        if r_idx == 0:
            continue
        for c_idx, cell in enumerate(ws2[r_idx + 1], 1):
            val = str(cell.value)
            if val == "-":
                cell.fill = fill_gray
            elif val.startswith("+"):
                try:
                    level = int(re.search(r"\d+", val).group())
                    if level <= 14:
                        cell.fill = fill_red
                    elif 15 <= level <= 16:
                        cell.fill = fill_yellow
                    elif level >= 17:
                        cell.fill = fill_green
                except:
                    cell.fill = fill_gray

    wb.save("result.xlsx")
    log("✅ 已保存 result.xlsx（明细 + 限时总览）")

# 写日志
with open("抓取日志.txt", "w", encoding="utf-8") as f:
    for line in log_entries:
        f.write(line + "\n")
log("📄 抓取日志已写入 抓取日志.txt")
