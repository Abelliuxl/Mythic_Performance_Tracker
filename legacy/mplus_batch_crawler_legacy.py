import pandas as pd
import re
import time
from urllib.parse import quote
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# ===== 副本映射与限时阈值（秒） =====
dungeon_name_map = {
    "Ara-Kara, City of Echoes": "艾拉-卡拉，回响之城",
    "Eco-Dome Al'dani": "奥尔达尼生态圆顶",
    "Halls of Atonement": "赎罪大厅",
    "Operation: Floodgate": "水闸行动",
    "Priory of the Sacred Flame": "圣焰隐修院",
    "Tazavesh: So'leah's Gambit": "塔扎维什: 索·莉亚的宏图",
    "Tazavesh: Streets of Wonder": "塔扎维什: 琳彩天街",
    "The Dawnbreaker": "破晨号",
}


dungeon_time_limit = {
    "艾拉-卡拉，回响之城": 30 * 60,          # 30:00
    "奥尔达尼生态圆顶": 31 * 60,              # 31:00
    "赎罪大厅": 31 * 60,                     # 31:00
    "水闸行动": 33 * 60,                     # 33:00
    "圣焰隐修院": 32 * 60 + 30,              # 32:30
    "塔扎维什: 索·莉亚的宏图": 30 * 60,        # 30:00
    "塔扎维什: 琳彩天街": 35 * 60,             # 35:00
    "破晨号": 31 * 60,                      # 31:00
}


# ===== 日志系统 =====
log_entries = []
def log(msg):
    print(msg)
    log_entries.append(msg)

# ===== 浏览器配置 =====
def create_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    options.add_argument("--proxy-server=http://192.168.5.101:7890")

    # ✅ 加这一行，强制不要等页面加载完就返回
    options.page_load_strategy = 'none'

    service = Service(executable_path="D:/ducument/chromedriver-win64/chromedriver.exe")
    return webdriver.Chrome(service=service, options=options)


# ===== 构建 URL =====
def build_url(server, character_name):
    return f"https://www.warcraftlogs.com/character/cn/{quote(server.strip(), safe='')}/{quote(character_name.strip(), safe='')}?zone=45"

# ===== 单角色爬取（修改版本）=====
def scrape_character(driver, server, character_name):
    url = build_url(server, character_name)

    max_attempts = 3
    for attempt in range(1, max_attempts + 1):
        log(f"  ⏳ 正在加载（尝试第 {attempt} 次）：{url}")
        try:
            driver.get(url)
            time.sleep(6)  # 等待副本表格渲染完成

            try:
                driver.execute_script("window.stop();")  # 停止广告等资源加载
            except Exception as e:
                log(f"  [警告] 无法停止加载：{e}")

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            rows = soup.find_all("tr", {"role": "row"})

            if rows:
                log(f"  ✅ 表格成功获取，共 {len(rows)} 行")
                break  # 成功获取，跳出循环
            else:
                log(f"  ⚠ 第 {attempt} 次未找到表格内容")

        except Exception as e:
            log(f"  [错误] 尝试第 {attempt} 次加载页面失败：{e}")

        # 若未成功，准备进行下一次循环（或最终放弃）

    else:
        log(f"[失败] 页面多次加载未成功，跳过角色：{character_name}")
        return []  # 所有尝试都失败，返回空记录

    # ✅ 成功获取 rows 后继续正常解析
    records = []
    for row in rows:
        try:
            a_tag = row.find("a", class_="Boss zone-boss-cell")
            if not a_tag:
                continue
            raw_dungeon = a_tag.text.strip()
            dungeon = dungeon_name_map.get(raw_dungeon, raw_dungeon)

            time_cells = row.find_all("td", class_="verbose main-table-number kills-cell")
            if len(time_cells) < 2:
                continue

            time_text = time_cells[1].get_text(strip=True)
            time_match = re.search(r"\d{1,2}:\d{2}", time_text)
            time_str = time_match.group(0) if time_match else "未知"
            level_match = re.search(r"\+(\d+)", time_text)
            plus_level = int(level_match.group(1)) if level_match else None

            t_split = time_str.split(":")
            run_seconds = int(t_split[0]) * 60 + int(t_split[1]) if len(t_split) == 2 else 9999
            limit = dungeon_time_limit.get(dungeon)
            on_time = (limit is not None and run_seconds <= limit)
            result = "是" if on_time else "否"

            records.append({
                "副本": dungeon,
                "限时层数": plus_level,
                "通关时间": time_str,
                "是否限时": result
            })
        except Exception as e:
            log(f"[错误] 解析副本行失败：{e}")
            continue

    return records


# ===== 主程序入口 =====
try:
    char_df = pd.read_excel("character_info.xlsx")
    log("✅ 成功读取 character_info.xlsx")
except Exception as e:
    log(f"❌ 无法读取角色文件：{e}")
    exit()

all_records = []

for _, row in char_df.iterrows():
    player = str(row["玩家"]).strip()
    name = str(row["角色名"]).strip()
    server = str(row["服务器"]).strip()

    log(f"\n—— 开始抓取：{player} / {name}（{server}）")
    driver = create_driver()
    try:
        char_data = scrape_character(driver, server, name)
        log(f"→ 获取成功，共 {len(char_data)} 条记录")
        for entry in char_data:
            entry.update({"玩家": player, "角色名": name, "服务器": server})
            all_records.append(entry)
    except Exception as e:
        log(f"⚠ 抓取失败：{name}：{e}")
    finally:
        driver.quit()

if not all_records:
    log("❌ 没有任何副本数据被抓取，终止写入。")
else:
    df = pd.DataFrame(all_records)
    df = df[["玩家", "角色名", "服务器", "副本", "通关时间", "限时层数", "是否限时"]]

    def format_level(row):
        lvl = row["限时层数"]
        if pd.isna(lvl):
            return "-"
        return f"+{int(lvl)}" if row["是否限时"] == "是" else f"+{int(lvl)}*"
    df["显示层数"] = df.apply(format_level, axis=1)

    pivot_df = df.pivot_table(
        index=["玩家", "角色名"],
        columns="副本",
        values="显示层数",
        aggfunc="first"
    ).fillna("-").reset_index()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "明细"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet("限时总览")

    from openpyxl.styles import PatternFill, Alignment

    fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    '''
    # ✅ 2-19 层独立颜色（彩虹风格，清晰区分）
    layer_color_map = {
        2:  "F44336", 3:  "E91E63", 4:  "9C27B0", 5:  "673AB7",
        6:  "3F51B5", 7:  "2196F3", 8:  "03A9F4", 9:  "00BCD4",
        10: "009688", 11: "4CAF50", 12: "8BC34A", 13: "CDDC39",
        14: "FFEB3B", 15: "FFC107", 16: "FF9800", 17: "FF5722",
        18: "795548", 19: "607D8B"
    }
    '''

    # 彩虹风格渐变色（蓝→绿→黄→橙→粉→红）
    layer_color_map = {
        2:  "E3F2FD",  3:  "BBDEFB",  4:  "81D4FA",  5:  "4DD0E1",
        6:  "4DB6AC",  7:  "81C784",  8:  "AED581",  9:  "DCE775",
        10: "FFF176", 11: "FFD54F", 12: "FFB74D", 13: "FF8A65",
        14: "F48FB1", 15: "F06292", 16: "EC407A", 17: "E91E63",
        18: "F44336", 19: "D32F2F",
    }

    # ✅ 写入并居中
    for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True), start=1):
        ws2.append(row)
        for c_idx, value in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="center", vertical="center")

    '''
    # ✅ 层数颜色标注
    for r_idx, row in enumerate(pivot_df.values, start=2):
        for c_idx, val in enumerate(row[2:], start=3):
            cell = ws2.cell(row=r_idx, column=c_idx)
            val_str = str(cell.value)
            if val_str == "-":
                cell.fill = fill_gray
            elif val_str.startswith("+"):
                try:
                    level = int(re.search(r"\d+", val_str).group())
                    hex_color = layer_color_map.get(level)
                    if hex_color:
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                except:
                    cell.fill = fill_gray
    '''

    # ✅ 层数颜色标注
    for r_idx, row in enumerate(pivot_df.values, start=2):
        for c_idx, val in enumerate(row[2:], start=3):
            cell = ws2.cell(row=r_idx, column=c_idx)
            val_str = str(cell.value)
            if val_str == "-":
                cell.fill = fill_gray
            elif val_str.startswith("+"):
                try:
                    level = int(re.search(r"\d+", val_str).group())
                    hex_color = layer_color_map.get(level)
                    if hex_color:
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                except:
                    cell.fill = fill_gray

    # ✅ 职业颜色映射并染色角色名列（第2列）
    class_color_map = {
        "死亡骑士": "C41F3B", "恶魔猎手": "A330C9", "德鲁伊": "FF7D0A", "猎人": "ABD473",
        "法师": "69CCF0", "武僧": "00FF96", "圣骑士": "F58CBA", "骑士": "F58CBA",
        "牧师": "FFFFFF", "潜行者": "FFF569", "盗贼": "FFF569", "萨满": "0070DE",
        "术士": "9482C9", "战士": "C79C6E"
    }
    char_class_map = dict(zip(char_df["角色名"], char_df["职业"]))

    for row in range(2, ws2.max_row + 1):
        char_name = ws2.cell(row=row, column=2).value
        char_class = char_class_map.get(char_name)
        if char_class:
            hex_color = class_color_map.get(char_class)
            if hex_color:
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                ws2.cell(row=row, column=2).fill = fill

    # ✅ 合并“玩家”列
    current_player = None
    start_row = 2
    for i in range(2, ws2.max_row + 2):
        if i <= ws2.max_row:
            value = ws2.cell(i, 1).value
        else:
            value = None

        if value != current_player:
            if current_player is not None and i - start_row > 1:
                ws2.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
                ws2.cell(start_row, 1).alignment = Alignment(horizontal="center", vertical="center")
            current_player = value
            start_row = i

    wb.save("result.xlsx")
    log("✅ 已保存 result.xlsx（明细 + 总览 + 层数颜色分布 + 职业染色）")




with open("抓取日志.txt", "w", encoding="utf-8") as f:
    for line in log_entries:
        f.write(line + "\n")
log("📄 抓取日志已写入 抓取日志.txt")
